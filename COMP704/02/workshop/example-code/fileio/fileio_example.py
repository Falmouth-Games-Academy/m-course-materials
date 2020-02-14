#https://www.w3schools.com/python/python_file_write.asp

f = open("demofile2.txt", "a")
f.write("Now the file has content!\n")
f.close()

#open and read the file after the appending:
f = open("demofile2.txt", "r")
print(f.read())
f.close()

#https://docs.python.org/3/library/csv.html
import csv
with open('eggs.csv', 'w', newline='') as csvfile:
    spamwriter = csv.writer(csvfile, delimiter=' ',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
    spamwriter.writerow(['Spam'] * 5 + ['Baked Beans'])
    spamwriter.writerow(['Spam', 'Lovely Spam', 'Wonderful Spam'])

with open('eggs.csv', newline='') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
    for row in spamreader:
        print(', '.join(row))

#https://openpyxl.readthedocs.io/en/stable/
#get openpyxl from the package manager
#preferences->project interpreter->package manager

#https://openpyxl.readthedocs.io/en/stable/tutorial.html

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")